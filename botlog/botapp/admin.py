from django.contrib import admin
from django.http import HttpResponse
from django.urls import reverse
from django.shortcuts import redirect
from django.contrib.auth.models import Group,User
from import_export import resources
from import_export.admin import ImportExportActionModelAdmin, ImportExportMixin
from .models import TelegramBotDialogModel,TelegramUserContactModel,UserContactModel
import openpyxl
from datetime import datetime as dt
from admin_extra_urls.api import button,ExtraUrlMixin
# Registered application's models here.

# button action to retrieve all data from all models
def export_excel_file():
    """ The function that creates excel document
    and returns HTTP file response 

    Args:
        model (_type_): _description_
        request (_type_): _description_
        queryset (_type_): _description_

    Returns:
        _type_: _description_
    """
    wh = openpyxl.Workbook()
    default_sheet = wh.active
    wh.remove(default_sheet)
    
    # telegram users
    tg_users_sheet = wh.create_sheet(title='TelegramContact')
    tg_users_sheet.append(['#','user_id','first_name','last_name','phone_number','created'])
    for user in TelegramUserContactModel.objects.all():
        tg_users_sheet.append([user.id,user.user_id,user.first_name,user.last_name,user.phone_number,user.created])
        
    # Dialogs
    dialog_sheet = wh.create_sheet(title='Dialogs ChatGPT')
    dialog_sheet.append(['#','username','prompt','response','created'])
    for dialog in TelegramBotDialogModel.objects.all():
        dialog_sheet.append([dialog.id,dialog.username,dialog.prompt,dialog.response,dialog.created])
    
    # frontend users
    f_users_sheet = wh.create_sheet(title='WebFormContacts')
    f_users_sheet.append(['#','name','email','created','phone','comment','contact'])
    for ct in UserContactModel.objects.all():
        f_users_sheet.append([ct.id,ct.name,ct.email,ct.created,ct.phone,ct.comment,ct.contact])
        
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="excel-data.xlsx"'
    wh.save(resp)
    return resp







# TELEGRAM USER CONTACT
class TelegramUserContactResource(resources.ModelResource):
    """TelegramUserContact Resource:
    * Import/export Exel option implementation class
    - This class extends admin panel with import/export Button

    Args:
        resources (_type_): _description_
    """
    class Meta:
        model = TelegramUserContactModel

class TelegramUserContactModelAdmin(ExtraUrlMixin,ImportExportActionModelAdmin,ImportExportMixin):
    list_display = ["user_id","first_name","last_name","phone_number","created"]
    list_filter = ["created"]
    ordering = ["created"]
    resource_classes = [TelegramUserContactResource]
    date_hierarchy = 'created'
    show_change_form_export = True
    @button(label="Все 100% в Excel")
    def get_data(self,request):
        return export_excel_file()


# USER CONTACT 

class UserContactResource(resources.ModelResource):
    """UserContactResource:
    * Import/export Exel option implementation class
    - This class extends admin panel with import/export Button
    Args:
        resources (_type_): _description_
    """
    class Meta:
        model = UserContactModel
        
class UserContactModelAdmin(ExtraUrlMixin,ImportExportActionModelAdmin,ImportExportMixin):
    list_display = ["name","email","phone","contact","created"]
    list_filter = ["created"]
    ordering = ["name","phone"]
    resource_classes = [UserContactResource]
    date_hierarchy = 'created'
    show_change_form_export = True
    @button(label="Все 100% в Excel")
    def get_data(self,request):
        return export_excel_file()
    
    
   


# DIALOG 
class TelegramBotDialogResource(resources.ModelResource):
    """TelegramBotDialogResourceResource
    * Import/export Exel option implementation class
    - This class extends admin panel with import/export Button
    Args:
        resources (_type_): _description_
    """
    class Meta:
        model = TelegramBotDialogModel
class TelegramBotDialogModelAdmin(ExtraUrlMixin,ImportExportActionModelAdmin,ImportExportMixin):
    """ TelegramBotDialogModelAdmin
    Args:
        ImportExportActionModelAdmin (_type_): _description_
        ImportExportMixin (_type_): _description_
    """
    list_display = ["username","prompt","created"]
    list_filter = ["created"]
    ordering = ["created"]
    resource_classes = [TelegramBotDialogResource]
    date_hierarchy = 'created'
    show_change_form_export = True
    @button(label="Все 100% в Excel")
    def get_data(self,request):
        return export_excel_file()
    
    
    
    
# Register on Admin Panel    
admin.site.register(TelegramUserContactModel,TelegramUserContactModelAdmin)
admin.site.register(UserContactModel,UserContactModelAdmin)
admin.site.register(TelegramBotDialogModel,TelegramBotDialogModelAdmin)



# Unregister
# admin.site.unregister(User)
# admin.site.unregister(Group)

# Django Admin settings: title | header | index text
admin.site.site_header = "BP Marketing BOT - администрирование"
admin.site.index_title = "BP - BusinessPad"


